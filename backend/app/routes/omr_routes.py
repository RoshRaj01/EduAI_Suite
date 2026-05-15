"""
OMR Routes — Computational grading endpoints.

All grading uses OpenCV + NumPy (OMRService).  No AI / LLM calls.
"""
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.omr import OMRJob, OMRSubmission
from app.services.omr_service import OMRService
from typing import List, Optional
import os
import io
import json
import hashlib
import tempfile

from PIL import Image
from io import BytesIO

omr_router = APIRouter(prefix="/api/omr", tags=["OMR Evaluation"])

# ── Helpers ─────────────────────────────────────────────────────────────

UPLOADS_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "uploads", "omr")
)


def _ensure_uploads():
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    return UPLOADS_DIR


def _save_temp(content: bytes, prefix: str = "temp") -> str:
    _ensure_uploads()
    name = f"{prefix}_{hashlib.md5(content).hexdigest()}.jpg"
    path = os.path.join(UPLOADS_DIR, name)
    with open(path, "wb") as f:
        f.write(content)
    return path


def _cleanup(path: str):
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def _score_answers(detected: dict, answer_key: dict) -> float:
    """Return percentage score."""
    if not answer_key:
        return 0.0
    total = len(answer_key)
    if total == 0:
        return 0.0
    correct = sum(
        1 for q, ans in answer_key.items()
        if detected.get(str(q)) == ans
    )
    return round((correct / total) * 100, 2)


# ── Job CRUD ────────────────────────────────────────────────────────────

@omr_router.post("/jobs")
async def create_omr_job(
    title: str = Form(...),
    answer_key: str = Form(None),
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    if not answer_key and not file:
        raise HTTPException(400, "Must provide either answer_key JSON or an image file")

    key_json = {}
    if file:
        content = await file.read()
        temp_path = _save_temp(content, "key")
        try:
            key_json = OMRService.process_omr_sheet(temp_path, num_questions=50)
            if not key_json:
                key_json = {str(i): "-" for i in range(1, 11)}
        except Exception as e:
            raise HTTPException(500, f"Failed to parse answer key image: {e}")
        finally:
            _cleanup(temp_path)
    else:
        try:
            key_json = json.loads(answer_key)
        except json.JSONDecodeError:
            raise HTTPException(400, "Invalid answer key JSON")

    job = OMRJob(title=title, answer_key=key_json)
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


@omr_router.get("/jobs")
def get_omr_jobs(db: Session = Depends(get_db)):
    return db.query(OMRJob).order_by(OMRJob.id.desc()).all()


@omr_router.get("/jobs/{job_id}")
def get_omr_job(job_id: int, db: Session = Depends(get_db)):
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    return job


@omr_router.delete("/jobs/{job_id}")
def delete_omr_job(job_id: int, db: Session = Depends(get_db)):
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).delete()
    db.delete(job)
    db.commit()
    return {"message": "Job deleted successfully"}


@omr_router.put("/jobs/{job_id}")
async def update_omr_job(
    job_id: int,
    title: str = Form(...),
    answer_key: str = Form(None),
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    job.title = title

    if file:
        content = await file.read()
        temp_path = _save_temp(content, "upd")
        try:
            key_json = OMRService.process_omr_sheet(temp_path, num_questions=50)
            if key_json:
                job.answer_key = key_json
        except Exception as e:
            print(f"Error updating with image: {e}")
        finally:
            _cleanup(temp_path)
    elif answer_key:
        try:
            job.answer_key = json.loads(answer_key)
        except json.JSONDecodeError:
            raise HTTPException(400, "Invalid answer key JSON")

    db.commit()
    db.refresh(job)
    return job


# ── Single-sheet upload (existing) ──────────────────────────────────────

@omr_router.post("/jobs/{job_id}/upload")
async def upload_omr_sheet(
    job_id: int,
    student_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    uploads = _ensure_uploads()
    file_path = os.path.join(uploads, file.filename)
    content = await file.read()
    with open(file_path, "wb") as buf:
        buf.write(content)

    image_url = f"/uploads/omr/{file.filename}"

    # Process with computational OMR
    try:
        detected_answers = OMRService.process_omr_sheet(
            file_path, num_questions=len(job.answer_key),
        )
        if not detected_answers:
            detected_answers = {k: "-" for k in job.answer_key.keys()}
    except Exception as e:
        print(f"OMR processing error: {e}")
        detected_answers = {k: "-" for k in job.answer_key.keys()}

    final_score = _score_answers(detected_answers, job.answer_key)

    submission = OMRSubmission(
        job_id=job_id,
        student_id=student_id,
        image_url=image_url,
        detected_answers=detected_answers,
        score=final_score,
        status="pending",
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)
    return submission


@omr_router.get("/jobs/{job_id}/submissions")
def get_omr_submissions(job_id: int, db: Session = Depends(get_db)):
    return db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).all()


@omr_router.put("/submissions/{sub_id}")
def update_submission(
    sub_id: int,
    score: float = Form(...),
    detected_answers: str = Form(...),
    db: Session = Depends(get_db),
):
    sub = db.query(OMRSubmission).filter(OMRSubmission.id == sub_id).first()
    if not sub:
        raise HTTPException(404, "Submission not found")

    sub.score = score
    try:
        sub.detected_answers = json.loads(detected_answers)
    except json.JSONDecodeError:
        pass
    sub.status = "verified"
    db.commit()
    db.refresh(sub)
    return sub


# ── Batch upload (NEW — Phase 3) ───────────────────────────────────────

@omr_router.post("/jobs/{job_id}/upload-batch")
async def upload_batch(
    job_id: int,
    student_ids: str = Form(""),
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """
    Upload multiple answer-sheet images (or a single PDF) at once.

    student_ids: comma-separated IDs matching each file, or empty
                 (auto-generates Sheet_1, Sheet_2, …).
    """
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    uploads = _ensure_uploads()
    id_list = [s.strip() for s in student_ids.split(",") if s.strip()]
    results = []

    image_paths = []

    # If single PDF, extract pages as images
    if len(files) == 1 and files[0].content_type == "application/pdf":
        pdf_bytes = await files[0].read()
        page_images = OMRService.extract_pages_from_pdf(pdf_bytes)
        for idx, img in enumerate(page_images):
            fname = f"batch_{job_id}_page_{idx + 1}.png"
            path = os.path.join(uploads, fname)
            cv2.imwrite(path, img)
            image_paths.append((path, fname))
    else:
        for f in files:
            content = await f.read()
            path = os.path.join(uploads, f.filename)
            with open(path, "wb") as buf:
                buf.write(content)
            image_paths.append((path, f.filename))

    num_q = len(job.answer_key)

    for idx, (file_path, fname) in enumerate(image_paths):
        sid = id_list[idx] if idx < len(id_list) else f"Sheet_{idx + 1}"
        image_url = f"/uploads/omr/{fname}"

        try:
            detected = OMRService.process_omr_sheet(file_path, num_questions=num_q)
            if not detected:
                detected = {k: "-" for k in job.answer_key.keys()}
        except Exception:
            detected = {k: "-" for k in job.answer_key.keys()}

        final_score = _score_answers(detected, job.answer_key)

        sub = OMRSubmission(
            job_id=job_id,
            student_id=sid,
            image_url=image_url,
            detected_answers=detected,
            score=final_score,
            status="pending",
        )
        db.add(sub)
        db.commit()
        db.refresh(sub)
        results.append({
            "id": sub.id,
            "student_id": sid,
            "score": final_score,
            "detected_answers": detected,
        })

    return {"success": True, "count": len(results), "submissions": results}


# We need cv2 for the batch PDF path
import cv2


# ── Uncertainties (NEW — Phase 3) ──────────────────────────────────────

@omr_router.get("/jobs/{job_id}/uncertainties")
def get_uncertainties(job_id: int, db: Session = Depends(get_db)):
    """List all uncertain answers across submissions for a job."""
    subs = db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).all()
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    uncertain_items = []
    for sub in subs:
        answers = sub.detected_answers or {}
        for q_num, ans in answers.items():
            if ans in ("?", "-"):
                uncertain_items.append({
                    "submission_id": sub.id,
                    "student_id": sub.student_id,
                    "question": q_num,
                    "detected": ans,
                    "correct": job.answer_key.get(str(q_num), "?"),
                })

    return {"success": True, "total": len(uncertain_items), "items": uncertain_items}


# ── Export: Excel (NEW — Phase 3) ──────────────────────────────────────

@omr_router.get("/jobs/{job_id}/export/excel")
def export_excel(job_id: int, db: Session = Depends(get_db)):
    """Download color-coded Excel with all submissions."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    subs = db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).all()
    if not subs:
        raise HTTPException(404, "No submissions found")

    answer_key = job.answer_key or {}
    q_nums = sorted(answer_key.keys(), key=lambda x: int(x))

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    fills = {
        "correct": PatternFill("solid", fgColor="C6EFCE"),
        "wrong": PatternFill("solid", fgColor="FFC7CE"),
        "uncertain": PatternFill("solid", fgColor="FFEB9C"),
        "excluded": PatternFill("solid", fgColor="D9D9D9"),
    }

    headers = ["Student ID"] + [f"Q{q}" for q in q_nums] + ["Score %", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for sub in subs:
        detected = sub.detected_answers or {}
        row = [sub.student_id]
        for q in q_nums:
            row.append(detected.get(str(q), "-"))
        row.append(sub.score)
        row.append(sub.status)
        ws.append(row)

        row_num = ws.max_row
        for col_idx, q in enumerate(q_nums):
            cell = ws.cell(row=row_num, column=col_idx + 2)
            ans = detected.get(str(q), "-")
            correct = answer_key.get(str(q), "")
            if correct == "X" or correct == "":
                cell.fill = fills["excluded"]
            elif ans == correct:
                cell.fill = fills["correct"]
            elif ans in ("?", "-"):
                cell.fill = fills["uncertain"]
            else:
                cell.fill = fills["wrong"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"OMR_Results_{job.title.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export: CSV (NEW — Phase 3) ────────────────────────────────────────

@omr_router.get("/jobs/{job_id}/export/csv")
def export_csv(job_id: int, db: Session = Depends(get_db)):
    """Download CSV of all submissions."""
    import csv as csv_mod

    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    subs = db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).all()
    if not subs:
        raise HTTPException(404, "No submissions found")

    answer_key = job.answer_key or {}
    q_nums = sorted(answer_key.keys(), key=lambda x: int(x))

    buf = io.StringIO()
    writer = csv_mod.writer(buf)

    headers = ["Student ID"] + [f"Q{q}" for q in q_nums] + ["Score %", "Status"]
    writer.writerow(headers)

    for sub in subs:
        detected = sub.detected_answers or {}
        row = [sub.student_id]
        for q in q_nums:
            ans = detected.get(str(q), "-")
            correct = answer_key.get(str(q), "")
            if ans == correct:
                row.append(f"[CORRECT] {ans}")
            elif ans in ("?", "-"):
                row.append(f"[UNCERTAIN] {ans}")
            elif correct in ("X", ""):
                row.append(f"[EXCLUDED] {ans}")
            else:
                row.append(f"[WRONG] {ans} (Key: {correct})")
        row.append(sub.score)
        row.append(sub.status)
        writer.writerow(row)

    csv_bytes = buf.getvalue().encode("utf-8")
    buf.close()

    filename = f"OMR_Results_{job.title.replace(' ', '_')}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Question Analytics (NEW — Phase 3) ─────────────────────────────────

@omr_router.get("/jobs/{job_id}/analytics")
def question_analytics(job_id: int, db: Session = Depends(get_db)):
    """Per-question performance breakdown."""
    job = db.query(OMRJob).filter(OMRJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")

    subs = db.query(OMRSubmission).filter(OMRSubmission.job_id == job_id).all()
    if not subs:
        return {"success": True, "total_students": 0, "analysis": []}

    answer_key = job.answer_key or {}
    total_students = len(subs)

    stats = {}
    for q_num, correct in answer_key.items():
        if correct in ("X", ""):
            continue
        stats[q_num] = {
            "question": int(q_num),
            "correct_answer": correct,
            "correct": 0,
            "wrong": 0,
            "uncertain": 0,
            "distribution": {},
        }

    for sub in subs:
        detected = sub.detected_answers or {}
        for q_num in stats:
            ans = detected.get(str(q_num), "-")
            if ans in ("?", "-"):
                stats[q_num]["uncertain"] += 1
            elif ans == stats[q_num]["correct_answer"]:
                stats[q_num]["correct"] += 1
            else:
                stats[q_num]["wrong"] += 1

            stats[q_num]["distribution"][ans] = (
                stats[q_num]["distribution"].get(ans, 0) + 1
            )

    analysis = []
    for q_num in sorted(stats, key=lambda x: int(x)):
        s = stats[q_num]
        answered = s["correct"] + s["wrong"]
        pct = round((s["correct"] / answered) * 100, 2) if answered > 0 else 0

        if pct >= 80:
            difficulty = "Easy"
        elif pct >= 50:
            difficulty = "Medium"
        else:
            difficulty = "Hard"

        analysis.append({
            "question": s["question"],
            "correct_answer": s["correct_answer"],
            "correct_count": s["correct"],
            "wrong_count": s["wrong"],
            "uncertain_count": s["uncertain"],
            "correct_percentage": pct,
            "difficulty": difficulty,
            "distribution": s["distribution"],
        })

    return {
        "success": True,
        "total_students": total_students,
        "analysis": analysis,
    }
